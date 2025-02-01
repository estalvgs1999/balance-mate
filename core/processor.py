import os
import shutil
import time

import pandas as pd
from openpyxl import load_workbook
from PyQt6.QtCore import QThread, pyqtSignal


def export_to_excel(dataframe, save_path):
    """
    Exports a pandas DataFrame to an Excel file.

    Args:
        dataframe (pd.DataFrame): The data to be exported.
        save_path (str): The path where the Excel file will be saved.

    Returns:
        tuple: (bool, str) indicating whether the export was successful and a message.
    """
    try:
        dataframe.to_excel(save_path, index=False)
        print("Exporting to Excel...")
        return True, "File exported successfully."
    except Exception as e:
        return False, f"Export error: {str(e)}"


class FileProcessor(QThread):
    """
    Class for processing CSV files and exporting data to Excel.

    Attributes:
        file_path (str): Path to the input CSV file.
        month (str): Report month.
        year (str): Report year.
    """

    progress = pyqtSignal(int)
    finished = pyqtSignal(bool, str, str)

    def __init__(self, file_path, month, year):
        super().__init__()
        self.file_path = file_path
        self.month = month
        self.year = year

    def run(self):
        """
        Executes the processing function and emits a signal upon completion.

        Emits:
            finished (bool, str): A signal indicating the success status and a
                                  message. True and a success message if processing
                                  is successful, False and an error message if an
                                  exception occurs.
        """
        try:
            success, output_file_path = self.process()
            self.finished.emit(
                success, "Proceso completado con éxito.", output_file_path
            )
        except Exception as e:
            self.finished.emit(False, f"Error durante el procesamiento: {str(e)}", None)

    def process(self):
        """
        Processes the CSV file, generates the report in Excel, and returns the process status.

        Returns:
            tuple: (bool, str) indicating whether the processing was successful and a message.
        """
        try:
            print("Starting data preparation...")
            outgoing_balance, incoming_balance = self.prepare_data(self.file_path)
            print("Data preparation completed.")
            print("Creating temporary Excel file...")
            temp_file_path = self.create_incoming_balance_excel(incoming_balance)
            print(f"Temporary Excel file created at {temp_file_path}")
            return True, temp_file_path
        except Exception as e:
            return False, f"Processing error: {str(e)}"

    def prepare_data(self, file_path):
        """
        Reads and processes the data from the CSV file.

        Args:
            file_path (str): Path to the CSV file.

        Returns:
            tuple: DataFrames for outgoing and incoming balances.
        """
        print("Reading file...")
        df = self.read_file(file_path)
        print("Cleaning data...")
        df = self.clean_data(df)
        print("Formatting numbers...")
        df = self.format_numbers(df)
        print("Renaming columns...")
        df = self.rename_columns(df)
        print("Sorting data by date...")
        df.sort_values(by="date", inplace=True)
        print("Resetting index...")
        df.reset_index(drop=True, inplace=True)

        df_out = df[df["debit"] > 0].copy()
        df_in = df[df["credit"] > 0].copy()

        return df_out, df_in

    def read_file(self, file_path):
        """
        Reads the CSV file.

        Args:
            file_path (str): Path to the CSV file.

        Returns:
            pd.DataFrame: Data from the CSV file.
        """
        print(f"Reading CSV file from {file_path}...")
        return pd.read_csv(file_path, sep=";", encoding="utf-8")

    def clean_data(self, df):
        """
        Cleans the data by removing unnecessary columns and invalid rows.

        Args:
            df (pd.DataFrame): Original DataFrame.

        Returns:
            pd.DataFrame: Cleaned DataFrame.
        """
        print("Dropping unnecessary columns and invalid rows...")
        df.drop(columns=["oficina"], inplace=True, errors="ignore")
        df = df[df["numeroDocumento"] != "TOTAL"]
        return df

    def format_numbers(self, df):
        """
        Converts numeric values from text to float.

        Args:
            df (pd.DataFrame): DataFrame with numeric data.

        Returns:
            pd.DataFrame: DataFrame with formatted numbers.
        """
        print("Converting numeric values...")
        df["debito"] = pd.to_numeric(
            df["debito"].str.replace(",", ""), errors="coerce"
        ).fillna(0)
        df["credito"] = pd.to_numeric(
            df["credito"].str.replace(",", ""), errors="coerce"
        ).fillna(0)
        return df

    def rename_columns(self, df):
        """
        Renames the DataFrame columns to standardize them.

        Args:
            df (pd.DataFrame): DataFrame with original columns.

        Returns:
            pd.DataFrame: DataFrame with renamed columns.
        """
        print("Renaming columns...")
        df.rename(
            columns={
                "fechaMovimiento": "date",
                "descripcion": "description",
                "debito": "debit",
                "credito": "credit",
                "numeroDocumento": "document_number",
            },
            inplace=True,
        )
        return df

    def create_incoming_balance_excel(self, incoming_balance):
        """
        Creates a temporary Excel file with the processed data.

        Args:
            incoming_balance (pd.DataFrame): Incoming balance data.

        Returns:
            str: Path to the generated Excel file.
        """
        temp_dir = os.path.join(os.path.dirname(os.path.dirname(__file__)), "temp")
        os.makedirs(temp_dir, exist_ok=True)

        timestamp = time.strftime("%Y%m%d-%H%M%S")
        temp_file_path = os.path.join(
            temp_dir, f"balance-mate-{self.month}-{self.year}-{timestamp}.xlsx"
        )

        shutil.copy(
            os.path.join("docs", "original", "balance-mate-template.xlsx"),
            temp_file_path,
        )

        print("Loading workbook...")
        wb = load_workbook(temp_file_path)
        ws = wb.active

        sheet_name = f"{self.month} {self.year}"
        ws.title = sheet_name
        start_row = 8
        for idx, row in incoming_balance.iterrows():
            ws[f"A{start_row}"] = row["description"]
            ws[f"B{start_row}"] = row["credit"]
            ws[f"C{start_row}"] = row["date"]
            ws[f"D{start_row}"] = row["document_number"]

            ws[f"B{start_row}"].number_format = "#,##0.00"
            start_row += 1
            self.progress.emit(int((idx + 1) / len(incoming_balance) * 100))
            time.sleep(0.02)

        print("Saving workbook...")
        wb.save(temp_file_path)

        return temp_file_path
